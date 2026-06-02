"""Generar Excel de backup con TODOS los resultados de pruebas gateway Mayo 11-27."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path('C:/Users/jorge/OneDrive/Documentos/FincaDiag_Modular')

# Load data
with open(BASE / 'consolidado_pi_may_11_27_FINAL.json', 'r', encoding='utf-8') as f:
    pi_data = json.load(f)

with open(BASE / 'gateway_test_results_11_20.json', 'r', encoding='utf-8') as f:
    win_11_20 = json.load(f)

with open(BASE / 'gateway_test_results_21_28.json', 'r', encoding='utf-8') as f:
    win_21_28 = json.load(f)

wb = Workbook()

# ============================================
# HOJA 1: RESUMEN POR VISITA (Pi + Windows)
# ============================================
ws1 = wb.active
ws1.title = "Resumen por Visita"

headers1 = [
    "Visita", "Fecha", "Sesiones motor", "Sesiones test Windows",
    "Pi-Schema", "Pi-TLS", "Pi-Resilience", "Pi-Subscribe", "Pi-Idempotency", "Pi-Obj4",
    "Win-Schema", "Win-Idempotency", "Win-Obj4",
    "Notas"
]

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Build visit map
visits_data = {}

# Motor sessions count from scan_all
from pathlib import Path as P2
data_dir = P2('C:/Users/jorge/OneDrive/Documentos/FincaDiag_Modular/data/processed/visits')
motor_counts = {}
for vd in sorted(data_dir.glob('Visita_*_05_2026')):
    day = int(vd.name.split('_')[1])
    if day < 11 or day > 28:
        continue
    count = 0
    for ses in vd.glob('sesiones/*/correlation_summary.json'):
        if 'BASELINE' in ses.parent.name:
            continue
        d = json.loads(ses.read_text())
        raw = d.get('matched_events')
        if raw is None: raw = d.get('matches', 0)
        matches = len(raw) if isinstance(raw, list) else (int(raw) if raw else 0)
        if matches > 0:
            count += 1
    motor_counts[day] = count

# Pi results
for r in pi_data['pruebas_raspberry_pi']['resultados']:
    visit = r['visit']
    day = int(visit.split('_')[1])
    visits_data[day] = {
        'visit': visit,
        'pi_schema': r['schema'],
        'pi_tls': r['tls'],
        'pi_resilience': r['resilience'],
        'pi_subscribe': r['subscribe'],
        'pi_idempotency': r['idempotency'],
        'pi_obj4': r['objective4'],
        'win_schema': 'N/A',
        'win_idempotency': 'N/A',
        'win_obj4': 'N/A',
        'win_sessions_tested': 0,
        'motor_sessions': motor_counts.get(day, 0),
        'notes': ''
    }

# Windows results 11-20
for r in win_11_20:
    day = int(r['visit'].split('_')[1])
    if day in visits_data:
        visits_data[day]['win_schema'] = 'PASS' if r['results']['schema']['pass'] else 'FAIL'
        visits_data[day]['win_idempotency'] = 'PASS' if r['results']['idempotency']['pass'] else 'FAIL'
        visits_data[day]['win_obj4'] = 'PASS' if r['results']['objective4']['pass'] else 'FAIL'
        visits_data[day]['win_sessions_tested'] += 1

# Windows results 21-28
for r in win_21_28:
    day = int(r['visit'].split('_')[1])
    if day in visits_data:
        visits_data[day]['win_schema'] = 'PASS' if r['schema'] else 'FAIL'
        visits_data[day]['win_idempotency'] = 'PASS' if r['idempotency'] else 'FAIL'
        visits_data[day]['win_obj4'] = 'PASS' if r['objective4'] else 'FAIL'
        visits_data[day]['win_sessions_tested'] += 1

# Fill sheet
row = 2
for day in sorted(visits_data):
    d = visits_data[day]
    data_row = [
        d['visit'], f"2026-05-{day:02d}", d['motor_sessions'], d['win_sessions_tested'],
        d['pi_schema'], d['pi_tls'], d['pi_resilience'], d['pi_subscribe'], d['pi_idempotency'], d['pi_obj4'],
        d['win_schema'], d['win_idempotency'], d['win_obj4'],
        d['notes']
    ]
    for col, val in enumerate(data_row, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if val == 'PASS':
            cell.fill = pass_fill
            cell.font = Font(bold=True, color='006100')
        elif val == 'FAIL':
            cell.fill = fail_fill
            cell.font = Font(bold=True, color='9C0006')
    row += 1

# Adjust column widths
widths = [14, 12, 14, 18, 10, 8, 12, 12, 14, 8, 12, 14, 10, 30]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Freeze header
ws1.freeze_panes = 'A2'

# ============================================
# HOJA 2: Sesiones Windows (Schema/Idem/Obj4)
# ============================================
ws2 = wb.create_sheet("Sesiones Windows")

headers2 = ["Visita", "Sesion", "Schema", "Idempotencia", "Obj4", "eta (%)", "Matches"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

all_win = []
# 11-20
for r in win_11_20:
    msg = r['results']['objective4']['msg']
    eta = ''
    if 'eta=' in msg:
        try: eta = str(round(float(msg.split('eta=')[1].split(',')[0]), 2))
        except: pass
    all_win.append({
        'visit': r['visit'],
        'session': r['session'],
        'schema': 'PASS' if r['results']['schema']['pass'] else 'FAIL',
        'idempotency': 'PASS' if r['results']['idempotency']['pass'] else 'FAIL',
        'obj4': 'PASS' if r['results']['objective4']['pass'] else 'FAIL',
        'eta': eta,
        'matches': r['matches']
    })

# 21-28
for r in win_21_28:
    msg = r.get('msg', '')
    eta = ''
    if 'eta=' in msg:
        try: eta = str(round(float(msg.split('eta=')[1].split(',')[0]), 2))
        except: pass
    all_win.append({
        'visit': r['visit'],
        'session': r['session'],
        'schema': 'PASS' if r['schema'] else 'FAIL',
        'idempotency': 'PASS' if r['idempotency'] else 'FAIL',
        'obj4': 'PASS' if r['objective4'] else 'FAIL',
        'eta': eta,
        'matches': r['matches']
    })

row = 2
for w in sorted(all_win, key=lambda x: (int(x['visit'].split('_')[1]), x['session'])):
    data_row = [w['visit'], w['session'], w['schema'], w['idempotency'], w['obj4'], w['eta'], w['matches']]
    for col, val in enumerate(data_row, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if val == 'PASS':
            cell.fill = pass_fill
            cell.font = Font(bold=True, color='006100')
        elif val == 'FAIL':
            cell.fill = fail_fill
            cell.font = Font(bold=True, color='9C0006')
    row += 1

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 45
for c in 'CDEFG':
    ws2.column_dimensions[c].width = 14
ws2.freeze_panes = 'A2'

# ============================================
# HOJA 3: Pruebas Raspberry Pi
# ============================================
ws3 = wb.create_sheet("Pruebas Raspberry Pi")

headers3 = ["Visita", "Fecha", "Sesion test Pi", "Schema", "TLS", "Resilience", "Subscribe", "Idempotency", "Obj4"]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

row = 2
for r in pi_data['pruebas_raspberry_pi']['resultados']:
    day = int(r['visit'].split('_')[1])
    session_name = r.get('session', 'N/A')
    data_row = [
        r['visit'], f"2026-05-{day:02d}", session_name,
        r['schema'], r['tls'], r['resilience'], r['subscribe'], r['idempotency'], r['objective4']
    ]
    for col, val in enumerate(data_row, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if val == 'PASS':
            cell.fill = pass_fill
            cell.font = Font(bold=True, color='006100')
        elif val == 'FAIL':
            cell.fill = fail_fill
            cell.font = Font(bold=True, color='9C0006')
    row += 1

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 45
for c in 'DEFGHI':
    ws3.column_dimensions[c].width = 14
ws3.freeze_panes = 'A2'

# ============================================
# HOJA 4: Resumen Ejecutivo
# ============================================
ws4 = wb.create_sheet("Resumen Ejecutivo")

ws4.cell(row=1, column=1, value="BACKUP RESULTADOS PRUEBAS GATEWAY").font = Font(bold=True, size=14)
ws4.cell(row=2, column=1, value="Período: 11-28 de mayo de 2026 (post-intervención)")
ws4.cell(row=3, column=1, value=f"Fecha generación: {pi_data['fecha_generacion']}")
ws4.cell(row=4, column=1, value="Nodo: gateway-esmeralda (Raspberry Pi)")
ws4.cell(row=5, column=1, value="")

ws4.cell(row=6, column=1, value="ESTADÍSTICAS GENERALES").font = Font(bold=True, size=12)
ws4.cell(row=7, column=1, value=f"Total visitas con pruebas Pi: {len(pi_data['pruebas_raspberry_pi']['resultados'])}")
ws4.cell(row=8, column=1, value=f"Total sesiones testeadas Windows: {len(all_win)}")
ws4.cell(row=9, column=1, value=f"Total sesiones motor con matches>0 (May 11-28): {sum(motor_counts.values())}")
ws4.cell(row=10, column=1, value="")

ws4.cell(row=11, column=1, value="RESULTADOS PI (17 visitas, Mayo 11-27)").font = Font(bold=True, size=12)
pi_results = pi_data['pruebas_raspberry_pi']['resultados']
ws4.cell(row=12, column=1, value=f"TLS Handshake: {sum(1 for r in pi_results if r['tls']=='PASS')}/{len(pi_results)} PASS")
ws4.cell(row=13, column=1, value=f"Resiliencia: {sum(1 for r in pi_results if r['resilience']=='PASS')}/{len(pi_results)} PASS")
ws4.cell(row=14, column=1, value=f"Suscripción MQTT: {sum(1 for r in pi_results if r['subscribe']=='PASS')}/{len(pi_results)} PASS")
ws4.cell(row=15, column=1, value=f"Schema: {sum(1 for r in pi_results if r['schema']=='PASS')}/{len(pi_results)} PASS")
ws4.cell(row=16, column=1, value=f"Objetivo 4: {sum(1 for r in pi_results if r['objective4']=='PASS')}/{len(pi_results)} PASS")
ws4.cell(row=17, column=1, value="")

ws4.cell(row=18, column=1, value="RESULTADOS WINDOWS (dry-run, 24 sesiones)").font = Font(bold=True, size=12)
ws4.cell(row=19, column=1, value=f"Schema JSON: {sum(1 for w in all_win if w['schema']=='PASS')}/{len(all_win)} PASS")
ws4.cell(row=20, column=1, value=f"Idempotencia: {sum(1 for w in all_win if w['idempotency']=='PASS')}/{len(all_win)} PASS")
ws4.cell(row=21, column=1, value=f"Consistencia η: {sum(1 for w in all_win if w['obj4']=='PASS')}/{len(all_win)} PASS")
etas = [float(w['eta']) for w in all_win if w['eta']]
if etas:
    ws4.cell(row=22, column=1, value=f"η̄ post-intervención: {round(sum(etas)/len(etas), 2)}% (rango: {min(etas)}-{max(etas)}%)")

ws4.cell(row=24, column=1, value="NOTAS IMPORTANTES").font = Font(bold=True, size=12)
ws4.cell(row=25, column=1, value="1. Pruebas Pi ejecutadas: 10 visitas (11-20) previo + 7 visitas (21-27) esta noche (29/05)")
ws4.cell(row=26, column=1, value="2. Pruebas Windows ejecutadas por dry-run sobre sesiones con correlation_summary.json")
ws4.cell(row=27, column=1, value="3. Los resultados 11-20 de Pi provienen del consolidado original; 21-27 corridos hoy")

for r in range(1, 28):
    ws4.cell(row=r, column=1).alignment = Alignment(vertical='center')

ws4.column_dimensions['A'].width = 80

# Save
excel_path = BASE / 'BACKUP_Resultados_Gateway_May11_28_2026.xlsx'
wb.save(str(excel_path))
print(f"Excel guardado: {excel_path}")
print(f"  - Hoja 1: Resumen por Visita ({len(visits_data)} filas)")
print(f"  - Hoja 2: Sesiones Windows ({len(all_win)} filas)")
print(f"  - Hoja 3: Pruebas Raspberry Pi ({len(pi_results)} filas)")
print(f"  - Hoja 4: Resumen Ejecutivo")
