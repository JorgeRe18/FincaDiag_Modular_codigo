"""Actualizar hoja Raspberry Pi en Excel con Obj4 corregido para 21-27."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path

BASE = Path(r'C:\Users\jorge\OneDrive\Documentos\FincaDiag_Modular')
xlsx_path = BASE / 'BACKUP_Suite_Completa_11_27_2026.xlsx'
wb = load_workbook(str(xlsx_path))

# Los 7 visitas que ahora tienen Obj4 PASS
fixed = {
    'Visita_21_05_2026': True,
    'Visita_22_05_2026': True,
    'Visita_23_05_2026': True,
    'Visita_24_05_2026': True,
    'Visita_25_05_2026': True,
    'Visita_26_05_2026': True,
    'Visita_27_05_2026': True,
}

ws = wb['Raspberry Pi 11-27']
pf = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ff = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Columna Obj4 es la 9 (I)
for row in range(2, ws.max_row + 1):
    visit = ws.cell(row=row, column=1).value
    if visit in fixed:
        cell = ws.cell(row=row, column=9)
        cell.value = 'PASS'
        cell.fill = pf
        cell.font = Font(bold=True, color='006100')

# Actualizar hoja Resumen tambien
ws3 = wb['Resumen']
# Contar PASS en columna I
pass_count = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=9).value == 'PASS')
total = ws.max_row - 1

ws3.cell(row=16, column=2, value=f"{pass_count}/{total}")

wb.save(str(xlsx_path))
print(f"Excel actualizado: {xlsx_path}")
print(f"Obj4 Pi: {pass_count}/{total} PASS")
