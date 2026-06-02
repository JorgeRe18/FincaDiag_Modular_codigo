"""Corregir hoja Sensibilidad Ventana para separar 38 sesiones vs 33 con matches."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

BASE = Path(r'C:\Users\jorge\OneDrive\Documentos\FincaDiag_Modular')
xlsx_path = BASE / 'RESULTADOS_FincaDiag_Caps4_5_6_Mayo2026.xlsx'
wb = load_workbook(str(xlsx_path))
ws = wb['Sensibilidad Ventana']

# Actualizar filas de resumen (1-5)
ws.cell(row=1, column=1, value="ANALISIS DE SENSIBILIDAD DE VENTANA DE CORRELACION")
ws.cell(row=1, column=1).font = Font(bold=True, size=13)

ws.cell(row=2, column=1, value="Periodo: Mayo 11-27, 2026 (post-intervencion)")

ws.cell(row=3, column=1, value="Sesiones post-intervencion con serial_events > 0: 38 (TODAS)")
ws.cell(row=3, column=1).font = Font(bold=True)

ws.cell(row=4, column=1, value="  - Con matches > 0 (gateway dry-run): 33 sesiones")
ws.cell(row=4, column=1).font = Font(color="006100")

ws.cell(row=5, column=1, value="  - Con matches = 0: 5 sesiones (analizadas igualmente)")
ws.cell(row=5, column=1).font = Font(color="9C5700")

# Los datos estan bien. Solo actualizamos las filas de promedio al final.
# Encontrar la fila del promedio
prom_row = None
for row in range(7, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and 'PROMEDIO' in str(val):
        prom_row = row
        break

if prom_row:
    ws.cell(row=prom_row, column=1, value=f"PROMEDIO (n=38, todas las sesiones post-intervencion)")
    ws.cell(row=prom_row, column=1).font = Font(bold=True)

    # Agregar fila de promedio SOLO con matches > 0 (n=33)
    # Calcular promedio de las filas con eta > 0
    etas250_match = []
    etas300_match = []
    for row in range(7, prom_row):
        eta250 = ws.cell(row=row, column=3).value
        eta300 = ws.cell(row=row, column=6).value
        if eta250 is not None and eta250 > 0:
            etas250_match.append(eta250)
            etas300_match.append(eta300)
    
    if etas250_match:
        media250_m = round(sum(etas250_match)/len(etas250_match), 2)
        media300_m = round(sum(etas300_match)/len(etas300_match), 2)
        
        new_row = prom_row + 1
        ws.insert_rows(new_row)
        ws.cell(row=new_row, column=1, value="PROMEDIO (n=33, solo sesiones con matches > 0)")
        ws.cell(row=new_row, column=1).font = Font(bold=True, color="006100")
        ws.cell(row=new_row, column=3, value=media250_m)
        ws.cell(row=new_row, column=6, value=media300_m)
        ws.cell(row=new_row, column=9, value=round(media300_m - media250_m, 2))
        
        # Copiar estilos de bordes del promedio anterior
        from openpyxl.styles import Border, Side
        bd = Border(*([Side(style='thin')]*4))
        for c in range(1, 11):
            ws.cell(row=new_row, column=c).border = bd
            ws.cell(row=new_row, column=c).alignment = Alignment(horizontal='center')

wb.save(str(xlsx_path))
print(f"Hoja 'Sensibilidad Ventana' actualizada.")
print(f"  38 sesiones post-intervencion (todas con serial_events > 0)")
print(f"  33 sesiones con matches > 0 (gateway dry-run)")
if etas250_match:
    print(f"  Promedio n=33: eta250={media250_m}%, eta300={media300_m}%")
