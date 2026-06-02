"""Extraer datos de eventos de red del Excel para figura eventos_red."""
import openpyxl
from pathlib import Path

BASE = Path(r'C:\Users\jorge\OneDrive\Documentos\FincaDiag_Modular')
EXCEL = BASE / 'RESULTADOS_FincaDiag_Caps4_5_6_Mayo2026.xlsx'

wb = openpyxl.load_workbook(EXCEL)

# Revisar hojas
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== Hoja: {name} ===")
    # Primeras 3 filas
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        print(row)
    print(f"... max_row={ws.max_row}, max_col={ws.max_column}")
