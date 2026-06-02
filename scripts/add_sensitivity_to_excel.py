"""Agregar hoja de sensibilidad al Excel de backup existente."""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path('C:/Users/jorge/OneDrive/Documentos/FincaDiag_Modular')
PROCESSED = BASE / 'data/processed/visits'

# Replicar logica de sensitivity_window.py para obtener datos
POST_START = (2026, 5, 11)

def parse_date(vn):
    parts = vn.replace('Visita_', '').split('_')
    return (int(parts[2]), int(parts[1]), int(parts[0]))

def simulate(matches, window_ms):
    matched = [m for m in matches if m['abs_delta_ms'] <= window_ms]
    serial_n = len(matches)
    n_matched = len(matched)
    eta = round(n_matched / serial_n * 100, 2) if serial_n else 0.0
    desfase = round(sum(m['abs_delta_ms'] for m in matched) / n_matched, 1) if matched else 0.0
    return n_matched, eta, desfase

rows = []
for vd in sorted(PROCESSED.iterdir()):
    if not vd.is_dir() or not vd.name.startswith('Visita_'):
        continue
    if parse_date(vd.name) < POST_START:
        continue
    for sp in sorted((vd / 'sesiones').iterdir()):
        if not sp.is_dir() or 'BASELINE_ONLY' in sp.name or 'Captura_' not in sp.name:
            continue
        corr_path = sp / 'correlation_summary.json'
        if not corr_path.exists():
            continue
        corr = json.loads(corr_path.read_text())
        serial_n = corr.get('serial_events', 0)
        matches = corr.get('matches', [])
        if serial_n == 0 or not matches:
            continue
        toma = 'AM' if 'AM__2AM' in sp.name else 'PM'
        date = vd.name.replace('Visita_','').replace('_2026','').replace('_05_','/')
        label = f"{date} {toma}"
        m250, eta250, d250 = simulate(matches, 250)
        m300, eta300, d300 = simulate(matches, 300)
        delta = round(eta300 - eta250, 2)
        nota = 'MEJORA' if delta > 0 else ''
        rows.append({
            'sesion': label, 'serial': serial_n,
            'eta250': eta250, 'm250': m250,
            'eta300': eta300, 'm300': m300,
            'delta': delta, 'nota': nota,
            'desfase250': d250, 'desfase300': d300
        })

n = len(rows)
media_250 = round(sum(r['eta250'] for r in rows)/n, 2) if n else 0
media_300 = round(sum(r['eta300'] for r in rows)/n, 2) if n else 0
mejoras = sum(1 for r in rows if r['delta'] > 0)

# Abrir Excel existente
xlsx_path = BASE / 'BACKUP_Suite_Completa_11_27_2026.xlsx'
wb = load_workbook(str(xlsx_path))

hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hfont = Font(bold=True, color="FFFFFF", size=11)
pf = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
bd = Border(*([Side(style='thin')]*4))

def style(cell, val, header=False):
    cell.border = bd
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if header:
        cell.fill = hf; cell.font = hfont
    elif val == 'MEJORA':
        cell.fill = pf; cell.font = Font(bold=True, color='006100')

ws = wb.create_sheet("Sensibilidad Ventana")

# Headers
heads = ["Sesion", "Serial Events", "eta 250ms (%)", "Matches 250", "Desfase 250 (ms)",
         "eta 300ms (%)", "Matches 300", "Desfase 300 (ms)", "Delta eta (pp)", "Nota"]
for col, h in enumerate(heads, 1):
    style(ws.cell(row=1, column=col, value=h), h, header=True)

# Data
for i, r in enumerate(rows, 2):
    vals = [r['sesion'], r['serial'], r['eta250'], r['m250'], r['desfase250'],
            r['eta300'], r['m300'], r['desfase300'], r['delta'], r['nota']]
    for col, v in enumerate(vals, 1):
        style(ws.cell(row=i, column=col, value=v), v)

# Promedio row
prom_row = n + 2
ws.cell(row=prom_row, column=1, value="PROMEDIO").font = Font(bold=True)
ws.cell(row=prom_row, column=3, value=media_250)
ws.cell(row=prom_row, column=6, value=media_300)
ws.cell(row=prom_row, column=9, value=round(media_300-media_250, 2))
for c in range(1, 11):
    ws.cell(row=prom_row, column=c).border = bd
    ws.cell(row=prom_row, column=c).alignment = Alignment(horizontal='center')

# Resumen arriba
ws.insert_rows(1, 5)
ws.cell(row=1, column=1, value="ANALISIS DE SENSIBILIDAD DE VENTANA DE CORRELACION").font = Font(bold=True, size=13)
ws.cell(row=2, column=1, value=f"Periodo: Mayo 11-27, 2026 (post-intervencion)")
ws.cell(row=3, column=1, value=f"Sesiones analizadas: {n}")
ws.cell(row=4, column=1, value=f"eta media W=250ms: {media_250}% | W=300ms: {media_300}% | Ganancia: +{round(media_300-media_250,2)} pp")
ws.cell(row=5, column=1, value=f"Sesiones con mejora al ampliar a 300ms: {mejoras}/{n}")

# Column widths
widths = [14, 13, 14, 12, 16, 14, 12, 16, 14, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(str(xlsx_path))
print(f"Hoja 'Sensibilidad Ventana' agregada a {xlsx_path}")
print(f"  Sesiones: {n}")
print(f"  eta 250ms: {media_250}%")
print(f"  eta 300ms: {media_300}%")
print(f"  Ganancia: +{round(media_300-media_250, 2)} pp")
print(f"  Mejoras: {mejoras}/{n}")
