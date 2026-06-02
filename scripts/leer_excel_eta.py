"""Leer datos eta del Excel consolidado."""
import openpyxl
from pathlib import Path

BASE = Path(r'C:\Users\jorge\OneDrive\Documentos\FincaDiag_Modular')
EXCEL = BASE / 'RESULTADOS_FincaDiag_Caps4_5_6_Mayo2026.xlsx'

wb = openpyxl.load_workbook(EXCEL)
print("Hojas disponibles:")
for name in wb.sheetnames:
    print(f"  - {name}")

# Leer hoja Sensibilidad Ventana completa
ws = wb['Sensibilidad Ventana']
print(f"\n--- Hoja: Sensibilidad Ventana ---")
print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    print(row)
    
# Extraer datos de eta
etas_250 = []
etas_300 = []
sesiones = []
for row in ws.iter_rows(min_row=6, values_only=True):  # saltar encabezados
    if row[0] is None or 'PROMEDIO' in str(row[0]):
        continue
    if row[2] is not None and isinstance(row[2], (int, float)):
        etas_250.append(float(row[2]))
        etas_300.append(float(row[5]) if row[5] is not None else float(row[2]))
        sesiones.append(row[0])

print(f"\nSesiones extraidas: {len(sesiones)}")
print(f"etas_250: {etas_250}")
print(f"etas_300: {etas_300}")
if etas_250:
    print(f"Media 250ms: {sum(etas_250)/len(etas_250):.4f}%")
    print(f"Media 300ms: {sum(etas_300)/len(etas_300):.4f}%")
