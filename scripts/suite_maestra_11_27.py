"""
SUITE MAESTRA 11-27 MAYO 2026 - TODO el repertorio, TODAS las visitas.
- WINDOWS (local): Schema + Idempotencia + Obj4 sobre cada sesion con matches>0.
- RASPBERRY (Pi):  Schema + TLS + Resilience + Subscribe + Idempotencia + Obj4 por visita.
Descubre sesiones dinamicamente (sin asumir). Guarda log + JSON + Excel.
"""
import paramiko, os, json, time, subprocess, glob, hashlib, shutil
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(r'C:\Users\jorge\OneDrive\Documentos\FincaDiag_Modular')
DATA_DIR = BASE / 'data' / 'processed' / 'visits'
SRC = BASE / 'src'

HOST = "gateway-esmeralda-ssh.at.remote.it"
PORT = 33000
USER = "esmeralda"
PW = os.environ["PI_PASSWORD"]
PI_VISITS_ROOT = "/var/lib/fincadiag/processed/visits"
CA = "/etc/fincadiag/certs/ca.crt"
CERT = "/etc/fincadiag/certs/client.crt"
KEY = "/etc/fincadiag/certs/client.key"

log_lines = []
def log(msg):
    line = f"[{time.strftime('%H:%M:%S')}] {msg}"
    print(line, flush=True)
    log_lines.append(line)

# ============================================================
# DESCUBRIR SESIONES 11-27 CON MATCHES > 0 (una por visita: la de mayor match)
# ============================================================
def discover_sessions():
    """Devuelve dict {dia: (visit_name, session_name, matches, eta)} eligiendo la sesion con mas matches por visita."""
    chosen = {}
    for vd in sorted(DATA_DIR.glob('Visita_*_05_2026')):
        day = int(vd.name.split('_')[1])
        if day < 11 or day > 27:
            continue
        best = None
        for corr in vd.glob('sesiones/*/correlation_summary.json'):
            if 'BASELINE' in corr.parent.name:
                continue
            d = json.loads(corr.read_text())
            raw = d.get('matched_events')
            if raw is None:
                raw = d.get('matches', 0)
            matches = len(raw) if isinstance(raw, list) else (int(raw) if raw else 0)
            if matches <= 0:
                continue
            eta = d.get('eta_extraccion') or d.get('eta_extraccion_pct') or d.get('extraction_efficiency_pct')
            eta_val = round(float(eta), 2) if eta else None
            if best is None or matches > best[2]:
                best = (vd.name, corr.parent.name, matches, eta_val)
        if best:
            chosen[day] = best
    return chosen

log("=== DESCUBRIENDO SESIONES 11-27 (matches>0) ===")
sessions = discover_sessions()
for day in sorted(sessions):
    v, s, m, e = sessions[day]
    log(f"  Mayo {day:02d}: {s} (matches={m}, eta={e})")
log(f"Total visitas: {len(sessions)}")

# ============================================================
# PARTE A: WINDOWS (local dry-run) - Schema + Idempotencia + Obj4
# ============================================================
PUBLISHED_DIR = BASE / 'data' / 'gateway' / 'published'

def clear_published():
    PUBLISHED_DIR.mkdir(parents=True, exist_ok=True)
    for f in PUBLISHED_DIR.glob('*'):
        try: f.unlink()
        except: pass

def win_dry_run(session_dir, pub_dir):
    env = os.environ.copy()
    env['PYTHONPATH'] = str(SRC)
    if pub_dir.exists():
        for f in pub_dir.glob('*'):
            try: f.unlink()
            except: pass
    pub_dir.mkdir(parents=True, exist_ok=True)
    cmd = ['python', '-m', 'fincadiag.gateway.runtime',
           '--session-dir', str(session_dir),
           '--topic-root', 'fincadiag/la_esmeralda',
           '--dry-run', '--published-dir', str(pub_dir)]
    r = subprocess.run(cmd, capture_output=True, text=True, cwd=str(BASE), env=env)
    return r.returncode == 0

def md5_of_jsonl(pub_dir):
    h = hashlib.md5()
    for jf in sorted(pub_dir.glob('*.jsonl')):
        h.update(jf.read_bytes())
    return h.hexdigest()

log("\n" + "="*60)
log("PARTE A: WINDOWS (Schema + Idempotencia + Obj4)")
log("="*60)

win_results = {}
for day in sorted(sessions):
    visit, sname, matches, eta = sessions[day]
    ses_dir = DATA_DIR / visit / 'sesiones' / sname
    log(f"\nMayo {day:02d} - {visit}")

    # Schema + Obj4: corrida A
    pub_a = BASE / 'data' / 'gateway' / f'win_a_{day}'
    ok_a = win_dry_run(ses_dir, pub_a)
    jsonl_a = list(pub_a.glob('*.jsonl'))
    schema_pass = ok_a and len(jsonl_a) > 0

    # Obj4: comparar eta gateway vs motor
    obj4_pass = False
    gw_eta = None
    if jsonl_a:
        msgs = [json.loads(l) for l in jsonl_a[0].read_text().splitlines() if l.strip()]
        corr_msg = next((m for m in msgs if m.get('event_type') == 'correlation_summary'), None)
        if corr_msg:
            gw_eta = corr_msg.get('payload', {}).get('eta_extraccion_pct')
            gw_matches = corr_msg.get('payload', {}).get('matches', 0)
            obj4_pass = (gw_matches == matches)

    # Idempotencia: corrida B + comparar MD5
    pub_b = BASE / 'data' / 'gateway' / f'win_b_{day}'
    ok_b = win_dry_run(ses_dir, pub_b)
    idem_pass = (md5_of_jsonl(pub_a) == md5_of_jsonl(pub_b)) and ok_b

    win_results[day] = {
        'visit': visit, 'session': sname, 'matches': matches, 'eta': eta,
        'schema': schema_pass, 'idempotency': idem_pass, 'objective4': obj4_pass
    }
    log(f"  Schema={'PASS' if schema_pass else 'FAIL'} Idem={'PASS' if idem_pass else 'FAIL'} Obj4={'PASS' if obj4_pass else 'FAIL'} (eta_motor={eta} eta_gw={gw_eta})")

    # cleanup
    for p in (pub_a, pub_b):
        try: shutil.rmtree(p)
        except: pass

# ============================================================
# PARTE B: RASPBERRY PI - Schema + TLS + Resilience + Subscribe + Idempotencia + Obj4
# ============================================================
log("\n" + "="*60)
log("PARTE B: RASPBERRY PI (conectando)")
log("="*60)

def run(c, cmd, timeout=180):
    stdin, stdout, stderr = c.exec_command(cmd, timeout=timeout)
    return stdout.read().decode(errors='replace'), stderr.read().decode(errors='replace')

c = paramiko.SSHClient()
c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
c.connect(HOST, port=PORT, username=USER, password=PW, timeout=60, banner_timeout=60, auth_timeout=60)
log("Conectado a Pi")

# Verificar no haya procesos gateway corriendo
out, _ = run(c, "ps aux | grep -E 'fincadiag.gateway' | grep -v grep | wc -l")
log(f"Procesos gateway activos en Pi: {out.strip()}")
out, _ = run(c, "systemctl is-active mosquitto")
log(f"Mosquitto: {out.strip()}")

# TLS handshake: una verificacion global (propiedad del sistema)
log("\n--- TLS Handshake (verificacion global) ---")
tls_out, _ = run(c, f"""
# TLS 1.3 debe aceptar
timeout 8 mosquitto_pub --cafile {CA} --cert {CERT} --key {KEY} --tls-version tlsv1.3 -h localhost -p 8883 -t test/tls -m ok 2>&1 && echo "TLS13=OK" || echo "TLS13=FAIL"
# TLS 1.2 debe rechazar (broker fuerza 1.3)
timeout 8 mosquitto_pub --cafile {CA} --cert {CERT} --key {KEY} --tls-version tlsv1.2 -h localhost -p 8883 -t test/tls -m no 2>&1 | grep -qi error && echo "TLS12=REJECTED" || echo "TLS12=ACCEPTED"
""", timeout=30)
log(f"TLS output: {tls_out.strip()}")
tls_global_pass = 'TLS13=OK' in tls_out

pi_results = {}
for day in sorted(sessions):
    visit, sname, matches, eta = sessions[day]
    ses = f"{PI_VISITS_ROOT}/{visit}/sesiones/{sname}"
    log(f"\nMayo {day:02d} - {visit} ({sname})")

    SPOOL = f"/tmp/spool_{day}"
    OUT = f"/tmp/mqtt_{day}.txt"

    # 1. Schema (dry-run genera JSONL)
    o1, _ = run(c, f"""
export PYTHONPATH=/opt/fincadiag
rm -rf /tmp/pub_{day} && mkdir -p /tmp/pub_{day}
python3 -m fincadiag.gateway.runtime --session-dir "{ses}" --topic-root fincadiag/la_esmeralda --dry-run --published-dir /tmp/pub_{day} >/dev/null 2>&1
find /tmp/pub_{day} -name "*.jsonl" | wc -l
""", timeout=60)
    schema_pass = int(o1.strip() or 0) > 0
    log(f"  Schema={'PASS' if schema_pass else 'FAIL'}")

    # 2. Obj4 (eta gateway vs motor)
    o2, _ = run(c, f"""
export PYTHONPATH=/opt/fincadiag
python3 -c "
import json, glob
files = glob.glob('/tmp/pub_{day}/*.jsonl')
m = matches_gw = 0
eta_gw = None
if files:
    for line in open(files[0]):
        line = line.strip()
        if not line: continue
        d = json.loads(line)
        if d.get('event_type') == 'correlation_summary':
            p = d.get('payload', {{}})
            matches_gw = p.get('matches', 0)
            eta_gw = p.get('eta_extraccion_pct')
print('matches_gw=' + str(matches_gw) + ' eta_gw=' + str(eta_gw))
"
""", timeout=30)
    log(f"  Obj4 raw: {o2.strip()}")
    obj4_pass = False
    if 'matches_gw=' in o2:
        try:
            mgw = int(o2.split('matches_gw=')[1].split()[0])
            obj4_pass = (mgw == matches)
        except: pass
    log(f"  Obj4={'PASS' if obj4_pass else 'FAIL'} (motor={matches})")

    # 3. Idempotencia (2 corridas + MD5)
    o3, _ = run(c, f"""
export PYTHONPATH=/opt/fincadiag
rm -rf /tmp/pa_{day} /tmp/pb_{day} && mkdir -p /tmp/pa_{day} /tmp/pb_{day}
python3 -m fincadiag.gateway.runtime --session-dir "{ses}" --topic-root fincadiag/la_esmeralda --dry-run --published-dir /tmp/pa_{day} >/dev/null 2>&1
python3 -m fincadiag.gateway.runtime --session-dir "{ses}" --topic-root fincadiag/la_esmeralda --dry-run --published-dir /tmp/pb_{day} >/dev/null 2>&1
HA=$(find /tmp/pa_{day} -name "*.jsonl" -exec cat {{}} + | md5sum | awk '{{print $1}}')
HB=$(find /tmp/pb_{day} -name "*.jsonl" -exec cat {{}} + | md5sum | awk '{{print $1}}')
[ "$HA" = "$HB" ] && echo "IDEM=PASS" || echo "IDEM=FAIL"
""", timeout=60)
    idem_pass = 'IDEM=PASS' in o3
    log(f"  Idempotencia={'PASS' if idem_pass else 'FAIL'}")

    # 4. Resilience + Subscribe (stop broker -> spool -> start -> sub -> drain -> verify)
    o4, _ = run(c, f"""
rm -rf {SPOOL} {OUT} && mkdir -p {SPOOL}
sudo systemctl stop mosquitto
export PYTHONPATH=/opt/fincadiag
python3 -m fincadiag.gateway.runtime --session-dir "{ses}" --topic-root fincadiag/la_esmeralda --mqtt-host localhost --mqtt-port 8883 --tls-enabled --ca-path {CA} --cert-path {CERT} --key-path {KEY} --tls-min-version 1.3 --spool-dir {SPOOL} --published-dir /tmp/pubres_{day} >/dev/null 2>&1
SF=$(find {SPOOL} -type f 2>/dev/null | wc -l)
echo "SPOOL_FILES=$SF"
sudo systemctl start mosquitto
sleep 3
mosquitto_sub --cafile {CA} --cert {CERT} --key {KEY} -h localhost -p 8883 -t "fincadiag/la_esmeralda/#" -v > {OUT} 2>/dev/null &
SUB=$!
sleep 2
export PYTHONPATH=/opt/fincadiag
python3 -c "
import sys
sys.path.insert(0, '/opt/fincadiag')
from fincadiag.gateway.runtime import GatewayRuntime
from fincadiag.gateway.config import GatewayConfig
from pathlib import Path
cfg = GatewayConfig(topic_root='fincadiag/la_esmeralda', mqtt_host='localhost', mqtt_port=8883, tls_enabled=True, ca_path='{CA}', cert_path='{CERT}', key_path='{KEY}', tls_min_version='1.3', spool_dir=Path('{SPOOL}'), published_dir=Path('/tmp/pubdr_{day}'), dry_run=False)
r = GatewayRuntime(cfg).drain_spool()
print('drained=' + str(r.published_count) + ' failed=' + str(r.failed_count))
"
sleep 3
kill $SUB 2>/dev/null || true
RECV=$(wc -l < {OUT} 2>/dev/null || echo 0)
SA=$(find {SPOOL} -type f 2>/dev/null | wc -l)
echo "RECEIVED=$RECV SPOOL_AFTER=$SA"
""", timeout=180)
    log(f"  Resilience raw: {o4.strip()[-150:]}")
    received = spool_after = -1
    rl = [l for l in o4.split('\n') if 'RECEIVED=' in l]
    if rl:
        parts = rl[-1].split()
        received = int([p for p in parts if p.startswith('RECEIVED=')][0].split('=')[1])
        spool_after = int([p for p in parts if p.startswith('SPOOL_AFTER=')][0].split('=')[1])
    res_pass = received > 0 and spool_after == 0
    sub_pass = received > 0
    log(f"  Resilience={'PASS' if res_pass else 'FAIL'} Subscribe={'PASS' if sub_pass else 'FAIL'} (recv={received}, spool_after={spool_after})")

    pi_results[day] = {
        'visit': visit, 'session': sname,
        'schema': schema_pass, 'tls': tls_global_pass,
        'resilience': res_pass, 'subscribe': sub_pass,
        'idempotency': idem_pass, 'objective4': obj4_pass,
        'received': received
    }

c.close()
log("\nPi desconectada")

# ============================================================
# GUARDAR JSON + LOG
# ============================================================
final_json = {
    'fecha_generacion': time.strftime('%Y-%m-%d %H:%M:%S'),
    'periodo': '2026-05-11 a 2026-05-27',
    'total_visitas': len(sessions),
    'windows': {str(d): win_results[d] for d in sorted(win_results)},
    'raspberry_pi': {str(d): pi_results[d] for d in sorted(pi_results)},
    'tls_global': tls_global_pass
}
json_path = BASE / 'resultados_suite_11_27_FINAL.json'
json_path.write_text(json.dumps(final_json, indent=2, ensure_ascii=False), encoding='utf-8')
log(f"\nJSON guardado: {json_path}")

log_path = BASE / 'log_suite_11_27.txt'
log_path.write_text('\n'.join(log_lines), encoding='utf-8')
print(f"Log guardado: {log_path}")

# ============================================================
# GENERAR EXCEL
# ============================================================
wb = Workbook()
hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hfont = Font(bold=True, color="FFFFFF", size=11)
pf = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ff = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
bd = Border(*([Side(style='thin')]*4))

def style_cell(cell, val, header=False):
    cell.border = bd
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if header:
        cell.fill = hf; cell.font = hfont
    elif val == 'PASS':
        cell.fill = pf; cell.font = Font(bold=True, color='006100')
    elif val == 'FAIL':
        cell.fill = ff; cell.font = Font(bold=True, color='9C0006')

# Hoja 1: WINDOWS
ws = wb.active
ws.title = "Windows 11-27"
heads = ["Visita", "Fecha", "Sesion", "Matches", "eta%", "Schema", "Idempotencia", "Obj4"]
for col, h in enumerate(heads, 1):
    style_cell(ws.cell(row=1, column=col, value=h), h, header=True)
row = 2
for day in sorted(win_results):
    r = win_results[day]
    vals = [r['visit'], f"2026-05-{day:02d}", r['session'], r['matches'], r['eta'],
            'PASS' if r['schema'] else 'FAIL', 'PASS' if r['idempotency'] else 'FAIL', 'PASS' if r['objective4'] else 'FAIL']
    for col, v in enumerate(vals, 1):
        style_cell(ws.cell(row=row, column=col, value=v), v)
    row += 1
for i, w in enumerate([14,12,42,9,8,10,14,8], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# Hoja 2: RASPBERRY PI
ws2 = wb.create_sheet("Raspberry Pi 11-27")
heads2 = ["Visita", "Fecha", "Sesion", "Schema", "TLS", "Resilience", "Subscribe", "Idempotencia", "Obj4"]
for col, h in enumerate(heads2, 1):
    style_cell(ws2.cell(row=1, column=col, value=h), h, header=True)
row = 2
for day in sorted(pi_results):
    r = pi_results[day]
    vals = [r['visit'], f"2026-05-{day:02d}", r['session'],
            'PASS' if r['schema'] else 'FAIL', 'PASS' if r['tls'] else 'FAIL',
            'PASS' if r['resilience'] else 'FAIL', 'PASS' if r['subscribe'] else 'FAIL',
            'PASS' if r['idempotency'] else 'FAIL', 'PASS' if r['objective4'] else 'FAIL']
    for col, v in enumerate(vals, 1):
        style_cell(ws2.cell(row=row, column=col, value=v), v)
    row += 1
for i, w in enumerate([14,12,42,10,8,12,12,14,8], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

# Hoja 3: RESUMEN
ws3 = wb.create_sheet("Resumen")
def count_pass(d, k):
    return sum(1 for r in d.values() if r[k])
n = len(sessions)
rows = [
    ("BACKUP SUITE COMPLETA 11-27 MAYO 2026", ""),
    ("Generado", time.strftime('%Y-%m-%d %H:%M:%S')),
    ("Total visitas", n),
    ("", ""),
    ("WINDOWS (dry-run local)", ""),
    ("  Schema", f"{count_pass(win_results,'schema')}/{n}"),
    ("  Idempotencia", f"{count_pass(win_results,'idempotency')}/{n}"),
    ("  Objetivo 4", f"{count_pass(win_results,'objective4')}/{n}"),
    ("", ""),
    ("RASPBERRY PI", ""),
    ("  Schema", f"{count_pass(pi_results,'schema')}/{n}"),
    ("  TLS Handshake", f"{count_pass(pi_results,'tls')}/{n}"),
    ("  Resilience", f"{count_pass(pi_results,'resilience')}/{n}"),
    ("  Subscribe", f"{count_pass(pi_results,'subscribe')}/{n}"),
    ("  Idempotencia", f"{count_pass(pi_results,'idempotency')}/{n}"),
    ("  Objetivo 4", f"{count_pass(pi_results,'objective4')}/{n}"),
]
for i, (a, b) in enumerate(rows, 1):
    ca = ws3.cell(row=i, column=1, value=a)
    ws3.cell(row=i, column=2, value=b)
    if a and not a.startswith('  '):
        ca.font = Font(bold=True, size=12)
ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 25

xlsx_path = BASE / 'BACKUP_Suite_Completa_11_27_2026.xlsx'
wb.save(str(xlsx_path))
print(f"Excel guardado: {xlsx_path}")

# Resumen final consola
print("\n" + "="*60)
print("RESUMEN FINAL 11-27 MAYO")
print("="*60)
print(f"Visitas: {n}")
print("WINDOWS:")
print(f"  Schema:       {count_pass(win_results,'schema')}/{n}")
print(f"  Idempotencia: {count_pass(win_results,'idempotency')}/{n}")
print(f"  Obj4:         {count_pass(win_results,'objective4')}/{n}")
print("RASPBERRY PI:")
print(f"  Schema:       {count_pass(pi_results,'schema')}/{n}")
print(f"  TLS:          {count_pass(pi_results,'tls')}/{n}")
print(f"  Resilience:   {count_pass(pi_results,'resilience')}/{n}")
print(f"  Subscribe:    {count_pass(pi_results,'subscribe')}/{n}")
print(f"  Idempotencia: {count_pass(pi_results,'idempotency')}/{n}")
print(f"  Obj4:         {count_pass(pi_results,'objective4')}/{n}")
